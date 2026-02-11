# Copyright (c) 2025-2026 Sebastián Muñoz
# Licensed under the MIT License

"""
Unit tests for data exchange infrastructure.

Tests the CSV/Excel handlers, field validation, and entity registry.
"""
# pyright: reportMissingImports=false

import io
from uuid import UUID, uuid4

import pytest

from app.domain.ports.data_exchange import (
    EntityConfig,
    EntityRegistry,
    ExportFormat,
    FieldConfig,
    FieldType,
    ImportMode,
    ReportFormat,
)
from app.infrastructure.data_exchange.csv_handler import CSVHandler
from app.infrastructure.data_exchange.excel_handler import (
    ExcelHandler,
    is_excel_available,
)

# ============================================================================
# FieldConfig Tests
# ============================================================================


class TestFieldConfig:
    """Tests for FieldConfig validation and transformation."""

    def test_required_field_validation_empty_value(self):
        """Required field should fail validation for empty values."""
        field = FieldConfig(
            name="email",
            display_name="Email",
            field_type=FieldType.EMAIL,
            required=True,
        )

        is_valid, error = field.validate("")
        assert not is_valid
        assert error is not None
        assert "required" in error.lower()

    def test_required_field_validation_none_value(self):
        """Required field should fail validation for None values."""
        field = FieldConfig(
            name="name",
            display_name="Name",
            field_type=FieldType.STRING,
            required=True,
        )

        is_valid, error = field.validate(None)
        assert not is_valid
        assert error is not None
        assert "required" in error.lower()

    def test_optional_field_validation_empty_value(self):
        """Optional field should pass validation for empty values."""
        field = FieldConfig(
            name="description",
            display_name="Description",
            field_type=FieldType.STRING,
            required=False,
        )

        is_valid, error = field.validate("")
        assert is_valid
        assert error is None

    def test_email_field_validation_valid(self):
        """Email field should accept valid email."""
        field = FieldConfig(
            name="email",
            display_name="Email",
            field_type=FieldType.EMAIL,
        )

        is_valid, error = field.validate("user@example.com")
        assert is_valid
        assert error is None

    def test_email_field_validation_invalid(self):
        """Email field should reject invalid email."""
        field = FieldConfig(
            name="email",
            display_name="Email",
            field_type=FieldType.EMAIL,
        )

        is_valid, error = field.validate("not-an-email")
        assert not is_valid
        assert error is not None
        assert "invalid" in error.lower()

    def test_integer_field_validation_valid(self):
        """Integer field should accept valid integers."""
        field = FieldConfig(
            name="age",
            display_name="Age",
            field_type=FieldType.INTEGER,
        )

        is_valid, error = field.validate(25)
        assert is_valid
        assert error is None

        is_valid, error = field.validate("42")
        assert is_valid
        assert error is None

    def test_integer_field_validation_invalid(self):
        """Integer field should reject invalid values."""
        field = FieldConfig(
            name="age",
            display_name="Age",
            field_type=FieldType.INTEGER,
        )

        is_valid, error = field.validate("not-a-number")
        assert not is_valid

    def test_boolean_field_validation_various_values(self):
        """Boolean field should accept various boolean representations."""
        field = FieldConfig(
            name="is_active",
            display_name="Is Active",
            field_type=FieldType.BOOLEAN,
        )

        valid_true_values = [True, "true", "True", "1", "yes", "sí", "si"]
        valid_false_values = [False, "false", "False", "0", "no"]

        for val in valid_true_values:
            is_valid, error = field.validate(val)
            assert is_valid, f"Failed for value: {val}"

        for val in valid_false_values:
            is_valid, error = field.validate(val)
            assert is_valid, f"Failed for value: {val}"

    def test_uuid_field_validation_valid(self):
        """UUID field should accept valid UUIDs."""
        field = FieldConfig(
            name="id",
            display_name="ID",
            field_type=FieldType.UUID,
        )

        test_uuid = str(uuid4())
        is_valid, error = field.validate(test_uuid)
        assert is_valid
        assert error is None

    def test_uuid_field_validation_invalid(self):
        """UUID field should reject invalid UUIDs."""
        field = FieldConfig(
            name="id",
            display_name="ID",
            field_type=FieldType.UUID,
        )

        is_valid, error = field.validate("not-a-uuid")
        assert not is_valid

    def test_choices_validation_valid(self):
        """Field with choices should accept valid choices."""
        field = FieldConfig(
            name="plan",
            display_name="Plan",
            field_type=FieldType.STRING,
            choices=["free", "starter", "professional"],
        )

        is_valid, error = field.validate("starter")
        assert is_valid
        assert error is None

    def test_choices_validation_invalid(self):
        """Field with choices should reject invalid choices."""
        field = FieldConfig(
            name="plan",
            display_name="Plan",
            field_type=FieldType.STRING,
            choices=["free", "starter", "professional"],
        )

        is_valid, error = field.validate("enterprise")
        assert not is_valid
        assert error is not None
        assert "invalid choice" in error.lower()

    def test_custom_validator(self):
        """Custom validator should be applied."""

        def validate_starts_with_a(value: str) -> bool:
            return value.lower().startswith("a")

        field = FieldConfig(
            name="name",
            display_name="Name",
            field_type=FieldType.STRING,
            validator=validate_starts_with_a,
        )

        is_valid, error = field.validate("Alice")
        assert is_valid

        is_valid, error = field.validate("Bob")
        assert not is_valid

    def test_transform_integer(self):
        """Transform should convert string to integer."""
        field = FieldConfig(
            name="count",
            display_name="Count",
            field_type=FieldType.INTEGER,
        )

        result = field.transform("42")
        assert result == 42
        assert isinstance(result, int)

    def test_transform_boolean(self):
        """Transform should convert string to boolean."""
        field = FieldConfig(
            name="is_active",
            display_name="Is Active",
            field_type=FieldType.BOOLEAN,
        )

        assert field.transform("true") is True
        assert field.transform("false") is False
        assert field.transform("1") is True
        assert field.transform("0") is False
        assert field.transform("yes") is True
        assert field.transform("no") is False

    def test_transform_uuid(self):
        """Transform should convert string to UUID."""
        field = FieldConfig(
            name="id",
            display_name="ID",
            field_type=FieldType.UUID,
        )

        test_uuid_str = "550e8400-e29b-41d4-a716-446655440000"
        result = field.transform(test_uuid_str)
        assert isinstance(result, UUID)
        assert str(result) == test_uuid_str

    def test_transform_with_default(self):
        """Transform should return default for empty values."""
        field = FieldConfig(
            name="status",
            display_name="Status",
            field_type=FieldType.STRING,
            default="pending",
        )

        result = field.transform("")
        assert result == "pending"

        result = field.transform(None)
        assert result == "pending"

    def test_transform_custom_transformer(self):
        """Custom transformer should be applied."""

        def uppercase_transform(value: str) -> str:
            return value.upper()

        field = FieldConfig(
            name="code",
            display_name="Code",
            field_type=FieldType.STRING,
            transformer=uppercase_transform,
        )

        result = field.transform("abc")
        assert result == "ABC"


# ============================================================================
# EntityRegistry Tests
# ============================================================================


class TestEntityRegistry:
    """Tests for EntityRegistry functionality."""

    def setup_method(self):
        """Clear registry before each test."""
        EntityRegistry.clear()

    def teardown_method(self):
        """Clean up after each test."""
        EntityRegistry.clear()

    def test_register_and_get_entity(self):
        """Should register and retrieve entity config."""
        config = EntityConfig(
            name="test_entity",
            display_name="Test Entity",
            model=object,  # Mock model
            fields=[
                FieldConfig(name="id", display_name="ID", field_type=FieldType.UUID),
            ],
            permission_resource="test_entity",
        )

        EntityRegistry.register(config)

        retrieved = EntityRegistry.get("test_entity")
        assert retrieved is not None
        assert retrieved.name == "test_entity"
        assert retrieved.display_name == "Test Entity"

    def test_get_nonexistent_entity(self):
        """Should return None for nonexistent entity."""
        result = EntityRegistry.get("nonexistent")
        assert result is None

    def test_list_all_entities(self):
        """Should list all registered entities."""
        configs = [
            EntityConfig(
                name=f"entity_{i}",
                display_name=f"Entity {i}",
                model=object,
                fields=[],
                permission_resource=f"entity_{i}",
            )
            for i in range(3)
        ]

        for config in configs:
            EntityRegistry.register(config)

        all_entities = EntityRegistry.list_all()
        assert len(all_entities) == 3

    def test_unregister_entity(self):
        """Should unregister entity."""
        config = EntityConfig(
            name="temp_entity",
            display_name="Temp Entity",
            model=object,
            fields=[],
            permission_resource="temp_entity",
        )

        EntityRegistry.register(config)
        assert EntityRegistry.get("temp_entity") is not None

        EntityRegistry.unregister("temp_entity")
        assert EntityRegistry.get("temp_entity") is None

    def test_list_exportable_entities(self):
        """Should list only entities with exportable fields."""
        exportable_config = EntityConfig(
            name="exportable",
            display_name="Exportable",
            model=object,
            fields=[
                FieldConfig(name="id", display_name="ID", exportable=True),
            ],
            permission_resource="exportable",
        )

        non_exportable_config = EntityConfig(
            name="non_exportable",
            display_name="Non Exportable",
            model=object,
            fields=[
                FieldConfig(name="id", display_name="ID", exportable=False),
            ],
            permission_resource="non_exportable",
        )

        EntityRegistry.register(exportable_config)
        EntityRegistry.register(non_exportable_config)

        exportable = EntityRegistry.list_exportable()
        assert len(exportable) == 1
        assert exportable[0].name == "exportable"

    def test_list_importable_entities(self):
        """Should list only entities with importable fields."""
        importable_config = EntityConfig(
            name="importable",
            display_name="Importable",
            model=object,
            fields=[
                FieldConfig(name="id", display_name="ID", importable=True),
            ],
            permission_resource="importable",
        )

        non_importable_config = EntityConfig(
            name="non_importable",
            display_name="Non Importable",
            model=object,
            fields=[
                FieldConfig(name="id", display_name="ID", importable=False),
            ],
            permission_resource="non_importable",
        )

        EntityRegistry.register(importable_config)
        EntityRegistry.register(non_importable_config)

        importable = EntityRegistry.list_importable()
        assert len(importable) == 1
        assert importable[0].name == "importable"


# ============================================================================
# EntityConfig Tests
# ============================================================================


class TestEntityConfig:
    """Tests for EntityConfig functionality."""

    def test_get_exportable_fields(self):
        """Should filter exportable fields."""
        config = EntityConfig(
            name="test",
            display_name="Test",
            model=object,
            fields=[
                FieldConfig(name="id", display_name="ID", exportable=True),
                FieldConfig(name="password", display_name="Password", exportable=False),
                FieldConfig(name="email", display_name="Email", exportable=True),
            ],
            permission_resource="test",
        )

        exportable = config.get_exportable_fields()
        assert len(exportable) == 2
        assert all(f.exportable for f in exportable)

    def test_get_importable_fields(self):
        """Should filter importable fields."""
        config = EntityConfig(
            name="test",
            display_name="Test",
            model=object,
            fields=[
                FieldConfig(name="id", display_name="ID", importable=False),
                FieldConfig(name="name", display_name="Name", importable=True),
                FieldConfig(name="email", display_name="Email", importable=True),
            ],
            permission_resource="test",
        )

        importable = config.get_importable_fields()
        assert len(importable) == 2
        assert all(f.importable for f in importable)

    def test_get_required_fields(self):
        """Should filter required importable fields."""
        config = EntityConfig(
            name="test",
            display_name="Test",
            model=object,
            fields=[
                FieldConfig(
                    name="id", display_name="ID", required=False, importable=False
                ),
                FieldConfig(
                    name="name", display_name="Name", required=True, importable=True
                ),
                FieldConfig(
                    name="email", display_name="Email", required=True, importable=True
                ),
                FieldConfig(
                    name="notes", display_name="Notes", required=False, importable=True
                ),
            ],
            permission_resource="test",
        )

        required = config.get_required_fields()
        assert len(required) == 2
        assert all(f.required and f.importable for f in required)

    def test_get_field_by_name(self):
        """Should get field by name."""
        config = EntityConfig(
            name="test",
            display_name="Test",
            model=object,
            fields=[
                FieldConfig(name="id", display_name="ID"),
                FieldConfig(name="email", display_name="Email"),
            ],
            permission_resource="test",
        )

        field = config.get_field("email")
        assert field is not None
        assert field.name == "email"

        not_found = config.get_field("nonexistent")
        assert not_found is None

    def test_report_title_default(self):
        """Should set default report title from display name."""
        config = EntityConfig(
            name="users",
            display_name="Users",
            model=object,
            fields=[],
            permission_resource="users",
        )

        assert "Users" in config.report_title


# ============================================================================
# CSVHandler Tests
# ============================================================================


class TestCSVHandler:
    """Tests for CSV reading and writing."""

    @pytest.fixture
    def simple_entity_config(self):
        """Create a simple entity config for testing."""
        return EntityConfig(
            name="test_entity",
            display_name="Test Entity",
            model=object,  # Mock model
            fields=[
                FieldConfig(
                    name="name",
                    display_name="Name",
                    field_type=FieldType.STRING,
                    required=True,
                    exportable=True,
                    importable=True,
                ),
                FieldConfig(
                    name="email",
                    display_name="Email",
                    field_type=FieldType.EMAIL,
                    required=False,
                    exportable=True,
                    importable=True,
                ),
                FieldConfig(
                    name="age",
                    display_name="Age",
                    field_type=FieldType.INTEGER,
                    required=False,
                    exportable=True,
                    importable=True,
                ),
            ],
            permission_resource="test_entity",
        )

    def test_read_csv_basic(self, simple_entity_config):
        """Should read basic CSV data."""
        csv_content = "Name,Email,Age\nJohn,john@test.com,30\nJane,jane@test.com,25\n"
        file_bytes = csv_content.encode("utf-8")

        handler = CSVHandler()
        rows = list(handler.read(file_bytes, simple_entity_config))

        assert len(rows) == 2
        # Each row is (row_num, row_data, errors)
        assert rows[0][1]["name"] == "John"
        assert rows[0][1]["email"] == "john@test.com"
        assert rows[1][1]["name"] == "Jane"

    def test_read_csv_with_quotes(self, simple_entity_config):
        """Should handle quoted values with commas."""
        csv_content = 'Name,Email,Age\n"John, Jr.",john@test.com,30\n'
        file_bytes = csv_content.encode("utf-8")

        handler = CSVHandler()
        rows = list(handler.read(file_bytes, simple_entity_config))

        assert len(rows) == 1
        assert rows[0][1]["name"] == "John, Jr."

    def test_read_csv_empty_file(self, simple_entity_config):
        """Should handle empty file gracefully."""
        csv_content = ""
        file_bytes = csv_content.encode("utf-8")

        handler = CSVHandler()
        rows = list(handler.read(file_bytes, simple_entity_config))

        assert len(rows) == 0

    def test_read_csv_header_only(self, simple_entity_config):
        """Should handle file with only headers."""
        csv_content = "Name,Email,Age\n"
        file_bytes = csv_content.encode("utf-8")

        handler = CSVHandler()
        rows = list(handler.read(file_bytes, simple_entity_config))

        assert len(rows) == 0

    def test_write_csv_basic(self, simple_entity_config):
        """Should write basic CSV data."""
        rows = [
            {"name": "John", "email": "john@test.com", "age": 30},
            {"name": "Jane", "email": "jane@test.com", "age": 25},
        ]

        handler = CSVHandler()
        content = handler.write(rows, simple_entity_config)

        csv_str = content.decode("utf-8")
        lines = csv_str.strip().split("\n")

        assert len(lines) == 3  # Header + 2 rows
        assert "Name" in lines[0]
        assert "Email" in lines[0]
        assert "John" in lines[1]

    def test_write_csv_empty_data(self, simple_entity_config):
        """Should write headers even with no data."""
        rows = []

        handler = CSVHandler()
        content = handler.write(rows, simple_entity_config)

        csv_str = content.decode("utf-8")
        lines = csv_str.strip().split("\n")

        assert len(lines) == 1  # Only header
        assert "Name" in lines[0]

    def test_generate_template(self, simple_entity_config):
        """Should generate template with headers and example row."""
        handler = CSVHandler()
        content = handler.generate_template(simple_entity_config)

        csv_str = content.decode("utf-8")
        lines = csv_str.strip().split("\n")

        assert len(lines) >= 1
        assert "Name" in lines[0]
        assert "Email" in lines[0]


# ============================================================================
# ExcelHandler Tests
# ============================================================================


class TestExcelHandler:
    """Tests for Excel reading and writing."""

    @pytest.mark.skipif(not is_excel_available(), reason="openpyxl not installed")
    def test_write_excel_basic(self):
        """Should write basic Excel data."""
        fields = [
            FieldConfig(name="name", display_name="Name"),
            FieldConfig(name="email", display_name="Email"),
        ]

        # Create a mock model for testing
        class MockModel:
            pass

        entity_config = EntityConfig(
            name="test_entity",
            display_name="Test Entity",
            model=MockModel,
            fields=fields,
            permission_resource="test:entity",
        )
        rows = [
            {"name": "John", "email": "john@test.com"},
            {"name": "Jane", "email": "jane@test.com"},
        ]

        handler = ExcelHandler()
        content = handler.write(rows, entity_config)

        # Verify it's valid Excel by reading it back
        from openpyxl import load_workbook  # pyright: ignore[reportMissingImports]

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        assert ws is not None
        assert ws["A1"].value == "Name"
        assert ws["B1"].value == "Email"
        assert ws["A2"].value == "John"

    @pytest.mark.skipif(not is_excel_available(), reason="openpyxl not installed")
    def test_read_excel_basic(self):
        """Should read basic Excel data."""
        from openpyxl import Workbook  # pyright: ignore[reportMissingImports]

        # Create field configs
        fields = [
            FieldConfig(name="name", display_name="Name"),
            FieldConfig(name="email", display_name="Email", field_type=FieldType.EMAIL),
        ]

        # Create a mock model for testing
        class MockModel:
            pass

        entity_config = EntityConfig(
            name="test_entity",
            display_name="Test Entity",
            model=MockModel,
            fields=fields,
            permission_resource="test:entity",
        )

        # Create a test workbook
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Name"
        ws["B1"] = "Email"
        ws["A2"] = "John"
        ws["B2"] = "john@test.com"
        ws["A3"] = "Jane"
        ws["B3"] = "jane@test.com"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()

        # Read with handler
        handler = ExcelHandler()
        rows = list(handler.read(content, entity_config))

        assert len(rows) == 2
        # Each row is (row_num, row_data, errors)
        assert rows[0][1]["name"] == "John"
        assert rows[0][1]["email"] == "john@test.com"

    def test_excel_availability_check(self):
        """Should correctly report Excel availability."""
        # This should not raise an error
        result = is_excel_available()
        assert isinstance(result, bool)


# ============================================================================
# Enum Tests
# ============================================================================


class TestEnums:
    """Tests for data exchange enums."""

    def test_field_type_values(self):
        """FieldType should have expected values."""
        assert FieldType.STRING.value == "string"
        assert FieldType.INTEGER.value == "integer"
        assert FieldType.BOOLEAN.value == "boolean"
        assert FieldType.EMAIL.value == "email"
        assert FieldType.UUID.value == "uuid"

    def test_import_mode_values(self):
        """ImportMode should have expected values."""
        assert ImportMode.INSERT.value == "insert"
        assert ImportMode.UPSERT.value == "upsert"
        assert ImportMode.UPDATE_ONLY.value == "update_only"

    def test_export_format_values(self):
        """ExportFormat should have expected values."""
        assert ExportFormat.CSV.value == "csv"
        assert ExportFormat.EXCEL.value == "excel"
        assert ExportFormat.JSON.value == "json"

    def test_report_format_values(self):
        """ReportFormat should have expected values."""
        assert ReportFormat.PDF.value == "pdf"
        assert ReportFormat.EXCEL.value == "excel"
        assert ReportFormat.CSV.value == "csv"
        assert ReportFormat.HTML.value == "html"
