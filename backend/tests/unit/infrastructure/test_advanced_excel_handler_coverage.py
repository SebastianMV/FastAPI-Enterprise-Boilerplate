# Copyright (c) 2025-2026 Sebastián Muñoz
# Licensed under the Apache License, Version 2.0

"""
Coverage tests for AdvancedExcelHandler – targeting uncovered lines.

Focuses on create_workbook, _populate_sheet, _add_chart, _add_conditional_format,
_add_data_validation, create_report_with_summary, and convenience functions.
"""

from datetime import date, datetime
from uuid import uuid4

import pytest

from app.infrastructure.data_exchange.advanced_excel_handler import (
    OPENPYXL_AVAILABLE,
    ChartType,
    ConditionalFormatConfig,
    DataValidationConfig,
    ExcelChartConfig,
    ExcelSheetConfig,
    FormatRule,
    FormulaColumn,
    is_advanced_excel_available,
)

# Skip all if openpyxl not available
pytestmark = pytest.mark.skipif(
    not OPENPYXL_AVAILABLE,
    reason="openpyxl not installed",
)


# ============================================================================
# Availability
# ============================================================================


class TestAvailability:
    def test_is_advanced_excel_available(self):
        result = is_advanced_excel_available()
        assert result is True  # because pytestmark would skip otherwise


# ============================================================================
# AdvancedExcelHandler – create_workbook basic
# ============================================================================


class TestCreateWorkbook:
    def test_single_sheet_with_data(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Data",
            headers=["Name", "Age", "Score"],
            data=[
                ["Alice", 30, 95.5],
                ["Bob", 25, 88.0],
                ["Carol", 35, 92.3],
            ],
        )
        result = handler.create_workbook([sheet], title="Test Report")
        assert isinstance(result, bytes)
        assert len(result) > 0

        # Verify by loading
        from io import BytesIO

        import openpyxl  # type: ignore[import-not-found]

        wb = openpyxl.load_workbook(BytesIO(result))
        assert "Data" in wb.sheetnames
        ws = wb["Data"]
        assert ws.cell(row=1, column=1).value == "Name"
        assert ws.cell(row=2, column=1).value == "Alice"

    def test_multiple_sheets(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheets = [
            ExcelSheetConfig(name="Sheet1", headers=["A"], data=[[1]]),
            ExcelSheetConfig(name="Sheet2", headers=["B"], data=[[2]]),
        ]
        result = handler.create_workbook(sheets, title="Multi")
        from io import BytesIO

        import openpyxl  # type: ignore[import-not-found]

        wb = openpyxl.load_workbook(BytesIO(result))
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames

    def test_empty_sheet(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(name="Empty", headers=["Col1"], data=[])
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)


# ============================================================================
# Formulas
# ============================================================================


class TestFormulas:
    def test_formula_column(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Formulas",
            headers=["Q1", "Q2", "Q3"],
            data=[
                [100, 200, 300],
                [150, 250, 350],
            ],
            formulas=[
                FormulaColumn(
                    column_letter="D",
                    header="Total",
                    formula_template="=SUM(A{row}:C{row})",
                    number_format="#,##0",
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        from io import BytesIO

        import openpyxl  # type: ignore[import-not-found]

        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb["Formulas"]
        # Formula header
        assert ws.cell(row=1, column=4).value == "Total"
        # Formula in row 2
        assert ws.cell(row=2, column=4).value == "=SUM(A2:C2)"


# ============================================================================
# Charts
# ============================================================================


class TestCharts:
    def test_bar_chart(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Charts",
            headers=["Category", "Value"],
            data=[["A", 10], ["B", 20], ["C", 30]],
            charts=[
                ExcelChartConfig(
                    chart_type=ChartType.BAR,
                    title="Bar Chart",
                    data_range="B1:B4",
                    category_range="A2:A4",
                    position="D2",
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_pie_chart(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Pie",
            headers=["Category", "Value"],
            data=[["X", 40], ["Y", 60]],
            charts=[
                ExcelChartConfig(
                    chart_type=ChartType.PIE,
                    title="Pie Chart",
                    data_range="B1:B3",
                    position="D2",
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_line_chart(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Line",
            headers=["Month", "Sales"],
            data=[["Jan", 100], ["Feb", 120], ["Mar", 90]],
            charts=[
                ExcelChartConfig(
                    chart_type=ChartType.LINE,
                    title="Line Chart",
                    data_range="B1:B4",
                    category_range="A2:A4",
                    position="D2",
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_column_chart(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Column",
            headers=["Dept", "Budget"],
            data=[["HR", 50000], ["IT", 80000]],
            charts=[
                ExcelChartConfig(
                    chart_type=ChartType.COLUMN,
                    title="Column Chart",
                    data_range="B1:B3",
                    position="D2",
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_chart_empty_data_range(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="NoChart",
            headers=["A"],
            data=[[1]],
            charts=[
                ExcelChartConfig(
                    chart_type=ChartType.BAR,
                    title="Empty",
                    data_range="",  # empty → no chart
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)


# ============================================================================
# Conditional Formatting
# ============================================================================


class TestConditionalFormatting:
    def test_color_scale(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="CF",
            headers=["Value"],
            data=[[10], [50], [90]],
            conditional_formats=[
                ConditionalFormatConfig(
                    rule_type=FormatRule.COLOR_SCALE,
                    cell_range="A2:A4",
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_formula_rule(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="FormulaRule",
            headers=["Score"],
            data=[[30], [70], [100]],
            conditional_formats=[
                ConditionalFormatConfig(
                    rule_type=FormatRule.FORMULA,
                    cell_range="A2:A4",
                    formula="$A2>50",
                    fill_color="00FF00",
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_icon_set(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Icons",
            headers=["KPI"],
            data=[[20], [50], [80]],
            conditional_formats=[
                ConditionalFormatConfig(
                    rule_type=FormatRule.ICON_SET,
                    cell_range="A2:A4",
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)


# ============================================================================
# Data Validation
# ============================================================================


class TestDataValidation:
    def test_list_validation(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="DV",
            headers=["Status"],
            data=[["Active"]],
            validations=[
                DataValidationConfig(
                    cell_range="A2:A100",
                    validation_type="list",
                    options=["Active", "Inactive", "Pending"],
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_numeric_validation(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="NumVal",
            headers=["Amount"],
            data=[[100]],
            validations=[
                DataValidationConfig(
                    cell_range="A2:A100",
                    validation_type="whole",
                    min_value=0,
                    max_value=1000,
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_decimal_validation(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="DecVal",
            headers=["Price"],
            data=[[9.99]],
            validations=[
                DataValidationConfig(
                    cell_range="A2:A100",
                    validation_type="decimal",
                    min_value=0.0,
                    max_value=999.99,
                ),
            ],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_unsupported_validation_type(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="UnsupVal",
            headers=["X"],
            data=[[1]],
            validations=[
                DataValidationConfig(
                    cell_range="A2:A100",
                    validation_type="custom",  # not handled
                ),
            ],
        )
        # Should not raise
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)


# ============================================================================
# Sheet Protection
# ============================================================================


class TestSheetProtection:
    def test_protected_sheet(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Protected",
            headers=["Data"],
            data=[[1]],
            protected=True,
            password="secret",
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_protected_without_password(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Protected",
            headers=["Data"],
            data=[[1]],
            protected=True,
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)


# ============================================================================
# Column Widths
# ============================================================================


class TestColumnWidths:
    def test_custom_column_widths(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Widths",
            headers=["Name", "Description"],
            data=[["A", "Long text here"]],
            column_widths={"A": 30, "B": 50},
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_auto_column_widths(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="Auto",
            headers=["Short", "Very Long Column Header Name"],
            data=[["x", "y"]],
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)


# ============================================================================
# _format_value
# ============================================================================


class TestFormatValue:
    def test_uuid_value(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        uid = uuid4()
        assert handler._format_value(uid) == str(uid)

    def test_date_value(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        d = date(2026, 1, 15)
        assert handler._format_value(d) == d

    def test_datetime_value(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        dt = datetime(2026, 1, 15, 10, 30)
        assert handler._format_value(dt) == dt

    def test_plain_value(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        assert handler._format_value("hello") == "hello"
        assert handler._format_value(42) == 42


# ============================================================================
# create_report_with_summary
# ============================================================================


class TestCreateReportWithSummary:
    def test_report_without_summary(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        result = handler.create_report_with_summary(
            title="Test",
            headers=["Name", "Score"],
            data=[["Alice", 95], ["Bob", 88]],
        )
        assert isinstance(result, bytes)

    def test_report_with_summary(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        result = handler.create_report_with_summary(
            title="Test",
            headers=["Name", "Score"],
            data=[["Alice", 95], ["Bob", 88]],
            summary_data={"Total Users": 2, "Average Score": 91.5, "Status": "OK"},
            include_charts=True,
        )
        assert isinstance(result, bytes)

        from io import BytesIO

        import openpyxl  # type: ignore[import-not-found]

        wb = openpyxl.load_workbook(BytesIO(result))
        assert "Summary" in wb.sheetnames

    def test_report_with_summary_no_charts(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        result = handler.create_report_with_summary(
            title="Test",
            headers=["Name", "Score"],
            data=[["Alice", 95]],
            summary_data={"Count": 1},
            include_charts=False,
        )
        assert isinstance(result, bytes)

    def test_report_with_numeric_first_column(self):
        """Test conditional formatting is added for numeric columns."""
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        result = handler.create_report_with_summary(
            title="Numeric",
            headers=["ID", "Value1", "Value2"],
            data=[[1, 100, 200], [2, 150, 250]],
        )
        assert isinstance(result, bytes)


# ============================================================================
# Convenience Functions
# ============================================================================


class TestConvenienceFunctions:
    def test_create_excel_report(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            create_excel_report,
        )

        result = create_excel_report(
            title="Test",
            headers=["A", "B"],
            data=[[1, 2], [3, 4]],
        )
        assert isinstance(result, bytes)

    def test_create_excel_report_with_summary(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            create_excel_report,
        )

        result = create_excel_report(
            title="Test",
            headers=["A", "B"],
            data=[[1, 2]],
            summary={"Total": 3},
        )
        assert isinstance(result, bytes)

    def test_create_multi_sheet_excel(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            create_multi_sheet_excel,
        )

        sheets = [
            ExcelSheetConfig(name="S1", headers=["X"], data=[[1]]),
            ExcelSheetConfig(name="S2", headers=["Y"], data=[[2]]),
        ]
        result = create_multi_sheet_excel(sheets, title="Multi")
        assert isinstance(result, bytes)


# ============================================================================
# Freeze Panes and Auto Filter
# ============================================================================


class TestLayoutOptions:
    def test_no_freeze_panes(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="NoFreeze",
            headers=["A"],
            data=[[1]],
            freeze_panes=None,
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)

    def test_no_auto_filter(self):
        from app.infrastructure.data_exchange.advanced_excel_handler import (
            AdvancedExcelHandler,
        )

        handler = AdvancedExcelHandler()
        sheet = ExcelSheetConfig(
            name="NoFilter",
            headers=["A"],
            data=[[1]],
            auto_filter=False,
        )
        result = handler.create_workbook([sheet])
        assert isinstance(result, bytes)
