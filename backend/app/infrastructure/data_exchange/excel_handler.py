# Copyright (c) 2025-2026 Sebastián Muñoz
# Licensed under the MIT License

"""
Excel Handler for data import/export.

Provides Excel reading and writing capabilities using openpyxl.
Falls back gracefully if openpyxl is not installed.
"""
# pyright: reportMissingImports=false, reportOptionalMemberAccess=false, reportOptionalCall=false

from __future__ import annotations

import io
from collections.abc import Iterator
from datetime import date, datetime
from typing import Any
from uuid import UUID

from app.domain.ports.data_exchange import EntityConfig, FieldConfig, FieldType

# Optional dependency - runtime import with graceful fallback
openpyxl: Any = None
Alignment: Any = None
Border: Any = None
Font: Any = None
PatternFill: Any = None
Side: Any = None
get_column_letter: Any = None

try:
    import openpyxl as _openpyxl
    from openpyxl.styles import (
        Alignment as _Alignment,
    )
    from openpyxl.styles import (
        Border as _Border,
    )
    from openpyxl.styles import (
        Font as _Font,
    )
    from openpyxl.styles import (
        PatternFill as _PatternFill,
    )
    from openpyxl.styles import (
        Side as _Side,
    )
    from openpyxl.utils import (
        get_column_letter as _get_column_letter,
    )

    openpyxl = _openpyxl
    Alignment = _Alignment
    Border = _Border
    Font = _Font
    PatternFill = _PatternFill
    Side = _Side
    get_column_letter = _get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None
    get_column_letter = None


class ExcelHandler:
    """
    Handler for Excel file operations.

    Provides methods for reading and writing Excel files
    with support for entity field configurations.

    Requires openpyxl to be installed.
    """

    def __init__(self) -> None:
        """Initialize Excel handler."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel support. "
                "Install it with: pip install openpyxl"
            )

    def read(
        self,
        file_content: bytes,
        entity_config: EntityConfig,
    ) -> Iterator[tuple[int, dict[str, Any], list[str]]]:
        """
        Read Excel file and yield rows with validation errors.

        Args:
            file_content: Excel file content as bytes
            entity_config: Entity configuration for field mapping

        Yields:
            Tuple of (row_number, row_data, errors)
        """
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = workbook.active

        if not sheet:
            return

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value else "")

        # Build field map
        field_map = self._build_field_map(headers, entity_config)

        # Read data rows
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            row_data = {}
            errors = []

            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers):
                    continue

                header = headers[col_idx]
                field_config = field_map.get(header)

                if not field_config:
                    continue

                if not field_config.importable:
                    continue

                value = cell.value

                # Validate
                is_valid, error_msg = field_config.validate(value)
                if not is_valid:
                    errors.append(error_msg or f"Invalid value for {header}")
                    continue

                # Transform
                try:
                    transformed = field_config.transform(value)
                    row_data[field_config.name] = transformed
                except Exception:
                    errors.append(f"Transform error for {header}")

            # Check required fields
            for field in entity_config.get_required_fields():
                if field.name not in row_data or row_data[field.name] is None:
                    errors.append(f"Missing required field: {field.display_name}")

            # Skip completely empty rows
            if not row_data and not errors:
                continue

            yield row_num, row_data, errors

    def write(
        self,
        data: list[dict[str, Any]],
        entity_config: EntityConfig,
        columns: list[str] | None = None,
        title: str | None = None,
    ) -> bytes:
        """
        Write data to Excel format.

        Args:
            data: List of dictionaries to export
            entity_config: Entity configuration for field mapping
            columns: Specific columns to include (None = all exportable)
            title: Optional title for the sheet

        Returns:
            Excel content as bytes
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = title or entity_config.display_name[:31]  # Excel limit

        # Get fields to export
        fields = entity_config.get_exportable_fields()
        if columns:
            fields = [f for f in fields if f.name in columns]

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write header
        for col_idx, field in enumerate(fields, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=field.display_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            # Set column width
            sheet.column_dimensions[get_column_letter(col_idx)].width = field.width

        # Write data
        for row_idx, row in enumerate(data, start=2):
            for col_idx, field in enumerate(fields, start=1):
                value = row.get(field.name)
                formatted = self._format_value(value, field)
                cell = sheet.cell(row=row_idx, column=col_idx, value=formatted)
                cell.border = thin_border

                # Apply number format if specified
                if field.format:
                    cell.number_format = field.format

        # Freeze header row
        sheet.freeze_panes = "A2"

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def generate_template(
        self,
        entity_config: EntityConfig,
        include_example: bool = True,
    ) -> bytes:
        """
        Generate an empty Excel template for import.

        Args:
            entity_config: Entity configuration
            include_example: Include an example row

        Returns:
            Template Excel content as bytes
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Importar"

        # Get importable fields
        fields = entity_config.get_importable_fields()

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        example_font = Font(italic=True, color="808080")

        # Write header
        for col_idx, field in enumerate(fields, start=1):
            header_text = field.display_name
            if field.required:
                header_text += " *"

            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            sheet.column_dimensions[get_column_letter(col_idx)].width = field.width

        # Write example row
        if include_example:
            for col_idx, field in enumerate(fields, start=1):
                example = self._get_example_value(field)
                cell = sheet.cell(row=2, column=col_idx, value=example)
                cell.font = example_font

        # Add instructions sheet
        instructions = workbook.create_sheet("Instrucciones")
        instructions.cell(
            row=1, column=1, value="Instrucciones de Importación"
        ).font = Font(bold=True, size=14)
        instructions.cell(
            row=3, column=1, value="1. Complete los datos en la hoja 'Importar'"
        )
        instructions.cell(
            row=4, column=1, value="2. Los campos marcados con * son obligatorios"
        )
        instructions.cell(
            row=5, column=1, value="3. Elimine la fila de ejemplo antes de importar"
        )
        instructions.cell(
            row=6, column=1, value="4. No modifique los encabezados de las columnas"
        )

        # Field descriptions
        instructions.cell(row=8, column=1, value="Campos disponibles:").font = Font(
            bold=True
        )
        for idx, field in enumerate(fields, start=9):
            req = "(Obligatorio)" if field.required else "(Opcional)"
            instructions.cell(row=idx, column=1, value=f"• {field.display_name} {req}")
            instructions.cell(
                row=idx, column=2, value=f"Tipo: {field.field_type.value}"
            )

        # Freeze header row
        sheet.freeze_panes = "A2"

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _build_field_map(
        self,
        headers: list[str],
        entity_config: EntityConfig,
    ) -> dict[str, FieldConfig]:
        """Build mapping from Excel headers to field configs."""
        field_map = {}

        for header in headers:
            if not header:
                continue

            # Remove required marker
            header_clean = header.replace(" *", "").strip()
            header_lower = header_clean.lower()

            for field in entity_config.fields:
                if field.display_name.lower() == header_lower:
                    field_map[header] = field
                    break
                if field.name.lower() == header_lower:
                    field_map[header] = field
                    break

        return field_map

    @staticmethod
    def _sanitize_formula(value: str) -> str:
        """Prevent Excel formula injection by prefixing dangerous chars with a single-quote."""
        if value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
            return "'" + value
        return value

    def _format_value(self, value: Any, field: FieldConfig) -> Any:
        """Format a value for Excel output."""
        if value is None:
            return ""

        if isinstance(value, (datetime, date)):
            return value  # Excel handles dates natively
        if isinstance(value, bool):
            return "Sí" if value else "No"
        if isinstance(value, UUID):
            return str(value)
        if isinstance(value, (dict, list)):
            import json

            return json.dumps(value, ensure_ascii=False)

        result = value
        if isinstance(result, str):
            result = self._sanitize_formula(result)
        return result

    def _get_example_value(self, field: FieldConfig) -> str:
        """Get an example value for a field."""
        examples: dict[FieldType, object] = {
            FieldType.STRING: "Texto ejemplo",
            FieldType.INTEGER: 123,
            FieldType.FLOAT: 123.45,
            FieldType.BOOLEAN: "Sí",
            FieldType.DATE: "2026-01-15",
            FieldType.DATETIME: "2026-01-15 10:30:00",
            FieldType.UUID: "550e8400-e29b-41d4-a716-446655440000",
            FieldType.EMAIL: "ejemplo@correo.com",
            FieldType.ENUM: field.choices[0] if field.choices else "opcion1",
            FieldType.JSON: '{"key": "value"}',
        }

        return str(examples.get(field.field_type, "valor"))


# Singleton instance
_excel_handler: ExcelHandler | None = None


def get_excel_handler() -> ExcelHandler:
    """
    Get Excel handler singleton.

    Raises:
        ImportError: If openpyxl is not installed
    """
    global _excel_handler
    if _excel_handler is None:
        _excel_handler = ExcelHandler()
    return _excel_handler


def is_excel_available() -> bool:
    """Check if Excel support is available."""
    return OPENPYXL_AVAILABLE
