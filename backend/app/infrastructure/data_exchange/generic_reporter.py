# Copyright (c) 2025-2026 Sebastián Muñoz
# Licensed under the Apache License, Version 2.0

"""
Generic Report Generator implementation.

Provides a flexible report generator that works with any registered entity.
"""
# pyright: reportMissingImports=false, reportOptionalMemberAccess=false, reportOptionalCall=false

from __future__ import annotations

import html
import io
import time
from datetime import UTC, datetime
from typing import Any
from uuid import UUID

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.domain.ports.data_exchange import (
    EntityConfig,
    EntityRegistry,
    FieldType,
    ReportFormat,
)
from app.domain.ports.reports import (
    ReportPort,
    ReportRequest,
    ReportResult,
    ReportSummary,
)
from app.infrastructure.data_exchange.csv_handler import get_csv_handler
from app.infrastructure.data_exchange.excel_handler import (
    is_excel_available,
)

# Type stubs for optional dependencies
from app.infrastructure.observability.logging import get_logger

logger = get_logger(__name__)


class GenericReporter(ReportPort):
    """
    Generic report generator that works with any registered entity.

    Uses the EntityRegistry to get field configurations and
    generates PDF, Excel, CSV, or HTML reports.
    """

    def __init__(self, session: AsyncSession):
        """
        Initialize the reporter.

        Args:
            session: SQLAlchemy async session
        """
        self.session = session

    def _apply_tenant_filter(
        self, query: Any, config: EntityConfig, tenant_id: UUID | None
    ) -> Any:
        """Apply tenant filter to query, raising error when tenant_id is missing for tenant-aware models."""
        if tenant_id and hasattr(config.model, "tenant_id"):
            return query.where(config.model.tenant_id == tenant_id)
        if not tenant_id and hasattr(config.model, "tenant_id"):
            raise ValueError("tenant_id is required for tenant-aware model reports")
        return query

    async def generate(self, request: ReportRequest) -> ReportResult:
        """
        Generate a report according to the request.

        Args:
            request: Report request configuration

        Returns:
            ReportResult with file content and metadata
        """
        start_time = time.time()

        # Get entity configuration
        config = EntityRegistry.get(request.entity)
        if not config:
            raise ValueError("Entity not found in registry")

        # Query data
        data = await self._query_data(config, request)
        rows = [self._model_to_dict(item, config) for item in data]

        # Get summary if requested
        summary = None
        if request.include_summary:
            summary = await self.get_summary(request)

        # Generate output based on format
        title = request.title or config.report_title

        if request.format == ReportFormat.CSV:
            content = get_csv_handler().write(rows, config, request.columns)
            content_type = "text/csv; charset=utf-8"
            extension = "csv"

        elif request.format == ReportFormat.EXCEL:
            content = self._generate_excel_report(
                rows, config, title, summary, request.columns
            )
            content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            extension = "xlsx"

        elif request.format == ReportFormat.HTML:
            content = self._generate_html_report(
                rows, config, title, summary, request.columns, request
            )
            content_type = "text/html; charset=utf-8"
            extension = "html"

        elif request.format == ReportFormat.PDF:
            # PDF requires additional dependency (weasyprint or similar)
            # For now, generate HTML that can be converted to PDF
            html_content = self._generate_html_report(
                rows, config, title, summary, request.columns, request
            )
            content = self._html_to_pdf(html_content)
            content_type = "application/pdf"
            extension = "pdf"

        else:
            raise ValueError("Unsupported report format")

        # Generate filename
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_{config.name}_{timestamp}.{extension}"

        duration_ms = (time.time() - start_time) * 1000

        return ReportResult(
            content=content,
            filename=filename,
            content_type=content_type,
            generated_at=datetime.now(UTC),
            row_count=len(rows),
            summary=summary,
            duration_ms=duration_ms,
        )

    async def get_preview(
        self,
        request: ReportRequest,
        limit: int = 10,
    ) -> list[dict[str, Any]]:
        """
        Get a preview of data before generating report.

        Args:
            request: Report request
            limit: Maximum rows to return

        Returns:
            List of dictionaries representing rows
        """
        config = EntityRegistry.get(request.entity)
        if not config:
            raise ValueError("Entity not found in registry")

        # Build query with limit
        query = select(config.model).limit(limit)

        # Apply tenant filter
        query = self._apply_tenant_filter(query, config, request.tenant_id)

        # Apply filters
        query = self._apply_filters(query, config, request)

        # Execute
        result = await self.session.execute(query)
        items = result.scalars().all()

        return [self._model_to_dict(item, config) for item in items]

    async def get_summary(
        self,
        request: ReportRequest,
    ) -> ReportSummary:
        """
        Get summary statistics without generating full report.

        Args:
            request: Report request

        Returns:
            ReportSummary with statistics
        """
        config = EntityRegistry.get(request.entity)
        if not config:
            raise ValueError("Entity not found in registry")

        # Get total count
        count_query = select(func.count()).select_from(config.model)

        count_query = self._apply_tenant_filter(count_query, config, request.tenant_id)

        result = await self.session.execute(count_query)
        total_records = result.scalar() or 0

        # Get grouped counts
        grouped_counts = {}
        if request.group_by:
            for group_field in request.group_by:
                if hasattr(config.model, group_field):
                    column = getattr(config.model, group_field)
                    group_query = (
                        select(column, func.count())
                        .select_from(config.model)
                        .group_by(column)
                    )

                    if request.tenant_id and hasattr(config.model, "tenant_id"):
                        group_query = group_query.where(
                            config.model.tenant_id == request.tenant_id
                        )

                    result = await self.session.execute(group_query)
                    for value, count in result.all():
                        key = (
                            f"{group_field}:{value}"
                            if value
                            else f"{group_field}:(vacío)"
                        )
                        grouped_counts[key] = count

        # Get numeric summaries
        numeric_summaries = {}
        for field in config.fields:
            if field.field_type in (FieldType.INTEGER, FieldType.FLOAT):
                if hasattr(config.model, field.name):
                    column = getattr(config.model, field.name)

                    agg_query = select(
                        func.min(column),
                        func.max(column),
                        func.avg(column),
                        func.sum(column),
                    ).select_from(config.model)

                    if request.tenant_id and hasattr(config.model, "tenant_id"):
                        agg_query = agg_query.where(
                            config.model.tenant_id == request.tenant_id
                        )

                    result = await self.session.execute(agg_query)
                    row = result.one_or_none()

                    if row and any(v is not None for v in row):
                        numeric_summaries[field.name] = {
                            "min": float(row[0]) if row[0] else 0,
                            "max": float(row[1]) if row[1] else 0,
                            "avg": float(row[2]) if row[2] else 0,
                            "sum": float(row[3]) if row[3] else 0,
                        }

        return ReportSummary(
            total_records=total_records,
            grouped_counts=grouped_counts,
            numeric_summaries=numeric_summaries,
        )

    async def _query_data(
        self,
        config: EntityConfig,
        request: ReportRequest,
    ) -> list[Any]:
        """Query data from the database."""
        query = select(config.model)

        # Exclude soft-deleted records
        if hasattr(config.model, "is_deleted"):
            query = query.where(config.model.is_deleted.is_(False))

        # Apply tenant filter
        query = self._apply_tenant_filter(query, config, request.tenant_id)

        # Apply filters
        query = self._apply_filters(query, config, request)

        # Apply date range — validate field against exportable fields
        allowed_fields = {f.name for f in config.fields}
        if (
            request.date_range_field
            and request.date_range_field in allowed_fields
            and hasattr(config.model, request.date_range_field)
        ):
            date_column = getattr(config.model, request.date_range_field)
            if request.date_from:
                query = query.where(date_column >= request.date_from)
            if request.date_to:
                query = query.where(date_column <= request.date_to)

        # Apply sorting
        sort_field = request.sort_by or config.default_sort
        if sort_field:
            descending = sort_field.startswith("-")
            if descending:
                sort_field = sort_field[1:]

            if (
                sort_field
                and sort_field in allowed_fields
                and hasattr(config.model, sort_field)
            ):
                column = getattr(config.model, sort_field)
                query = query.order_by(column.desc() if descending else column)

        # Apply limit
        query = query.limit(config.max_export_rows)

        # Execute
        result = await self.session.execute(query)
        return list(result.scalars().all())

    def _apply_filters(
        self,
        query: Any,
        config: EntityConfig,
        request: ReportRequest,
    ) -> Any:
        """Apply filters to a query."""
        for filter_item in request.filters:
            allowed_fields = {f.name for f in config.fields}
            if filter_item.field not in allowed_fields or not hasattr(
                config.model, filter_item.field
            ):
                continue

            column = getattr(config.model, filter_item.field)
            value = filter_item.value
            operator = filter_item.operator

            if operator == "eq":
                query = query.where(column == value)
            elif operator == "ne":
                query = query.where(column != value)
            elif operator == "gt":
                query = query.where(column > value)
            elif operator == "gte":
                query = query.where(column >= value)
            elif operator == "lt":
                query = query.where(column < value)
            elif operator == "lte":
                query = query.where(column <= value)
            elif operator == "in":
                query = query.where(column.in_(value))
            elif operator == "contains":
                escaped = (
                    str(value)
                    .replace("\\", "\\\\")
                    .replace("%", "\\%")
                    .replace("_", "\\_")
                )
                query = query.where(column.ilike(f"%{escaped}%", escape="\\"))

        return query

    def _model_to_dict(
        self,
        instance: Any,
        config: EntityConfig,
    ) -> dict[str, Any]:
        """Convert a model instance to a dictionary."""
        result = {}
        for field in config.get_exportable_fields():
            if hasattr(instance, field.name):
                result[field.name] = getattr(instance, field.name)
        return result

    def _generate_excel_report(
        self,
        data: list[dict[str, Any]],
        config: EntityConfig,
        title: str,
        summary: ReportSummary | None,
        columns: list[str] | None,
    ) -> bytes:
        """Generate an Excel report with formatting."""
        if not is_excel_available():
            raise ValueError("Excel support not available. Install openpyxl.")

        import openpyxl  # pyright: ignore[reportMissingModuleSource]
        from openpyxl.styles import (  # pyright: ignore[reportMissingModuleSource]
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import (
            get_column_letter,  # pyright: ignore[reportMissingModuleSource]
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Datos"

        # Get fields
        fields = config.get_exportable_fields()
        if columns:
            fields = [f for f in fields if f.name in columns]

        # Title
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(fields)
        )
        title_cell = sheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")

        # Subtitle with date
        sheet.merge_cells(
            start_row=2, start_column=1, end_row=2, end_column=len(fields)
        )
        subtitle_cell = sheet.cell(
            row=2,
            column=1,
            value=f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M')}",
        )
        subtitle_cell.font = Font(italic=True, size=10)
        subtitle_cell.alignment = Alignment(horizontal="center")

        # Headers
        header_row = 4
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, field in enumerate(fields, start=1):
            cell = sheet.cell(row=header_row, column=col_idx, value=field.display_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[get_column_letter(col_idx)].width = field.width

        # Data
        for row_idx, row in enumerate(data, start=header_row + 1):
            for col_idx, field in enumerate(fields, start=1):
                value = row.get(field.name)
                cell = sheet.cell(
                    row=row_idx, column=col_idx, value=self._format_cell_value(value)
                )
                cell.border = thin_border

        # Summary sheet if available
        if summary:
            summary_sheet = workbook.create_sheet("Resumen")
            summary_sheet.cell(
                row=1, column=1, value="Resumen del Reporte"
            ).font = Font(bold=True, size=14)
            summary_sheet.cell(row=3, column=1, value="Total de registros:")
            summary_sheet.cell(row=3, column=2, value=summary.total_records)

            if summary.grouped_counts:
                summary_sheet.cell(
                    row=5, column=1, value="Conteos por grupo:"
                ).font = Font(bold=True)
                summary_row = 6
                for key, count in summary.grouped_counts.items():
                    summary_sheet.cell(row=summary_row, column=1, value=key)
                    summary_sheet.cell(row=summary_row, column=2, value=count)
                    summary_row += 1

        # Freeze header
        sheet.freeze_panes = f"A{header_row + 1}"

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _generate_html_report(
        self,
        data: list[dict[str, Any]],
        config: EntityConfig,
        title: str,
        summary: ReportSummary | None,
        columns: list[str] | None,
        request: ReportRequest | None = None,
    ) -> bytes:
        """Generate an HTML report with professional headers/footers for PDF."""
        fields = config.get_exportable_fields()
        if columns:
            fields = [f for f in fields if f.name in columns]

        # Get header/footer options from request or use defaults
        header_text = title
        footer_text = "Confidencial"
        show_page_numbers = True
        company_name = None

        if request:
            header_text = request.header_text or title
            footer_text = request.footer_text
            show_page_numbers = request.show_page_numbers
            company_name = request.company_name

        # Build page number content
        page_number_content = (
            '"Página " counter(page) " de " counter(pages)'
            if show_page_numbers
            else '""'
        )

        def _escape_css_string(value: str | None) -> str:
            """Escape a value for safe interpolation into CSS content: strings."""
            if not value:
                return ""
            return (
                value.replace("\\", "\\\\")
                .replace('"', '\\"')
                .replace("\n", "\\A ")
                .replace("\r", "\\D ")
                .replace("\0", "")
            )

        # Build header left content (company name or empty)
        safe_css_company = _escape_css_string(company_name)
        safe_css_header = _escape_css_string(header_text)
        header_left = (
            f'"{safe_css_company}"' if company_name else f'"{safe_css_header}"'
        )

        # Build footer content
        generated_text = f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M')}"
        safe_css_footer = _escape_css_string(footer_text)
        footer_content = f'"{generated_text} | {safe_css_footer}"'

        html_out = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{html.escape(title)}</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            margin: 40px;
            color: #333;
        }}
        h1 {{
            color: #1a365d;
            border-bottom: 2px solid #4472C4;
            padding-bottom: 10px;
        }}
        .company-header {{
            color: #4472C4;
            font-size: 0.85em;
            margin-bottom: 5px;
            font-weight: 600;
        }}
        .meta {{
            color: #666;
            font-size: 0.9em;
            margin-bottom: 20px;
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
            margin-top: 20px;
        }}
        th {{
            background-color: #4472C4;
            color: white;
            padding: 12px 8px;
            text-align: left;
            font-weight: 600;
        }}
        td {{
            padding: 10px 8px;
            border-bottom: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background-color: #f8f9fa;
        }}
        tr:hover {{
            background-color: #e9ecef;
        }}
        .summary {{
            background-color: #f0f4f8;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .summary h2 {{
            margin-top: 0;
            color: #2d3748;
        }}
        .stat {{
            display: inline-block;
            margin-right: 30px;
        }}
        .stat-value {{
            font-size: 24px;
            font-weight: bold;
            color: #4472C4;
        }}
        .stat-label {{
            font-size: 12px;
            color: #666;
        }}
        /* PDF Page Layout with Headers/Footers (WeasyPrint compatible) */
        @page {{
            size: A4;
            margin: 2.5cm 1.5cm 2.5cm 1.5cm;
            @top-left {{
                content: {header_left};
                font-size: 10px;
                color: #666;
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            }}
            @top-right {{
                content: {page_number_content};
                font-size: 10px;
                color: #666;
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            }}
            @bottom-left {{
                content: {footer_content};
                font-size: 9px;
                color: #999;
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            }}
            @bottom-right {{
                content: "{safe_css_footer}";
                font-size: 9px;
                color: #999;
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            }}
        }}
        @page :first {{
            margin-top: 1.5cm;
            @top-left {{ content: none; }}
            @top-right {{ content: none; }}
        }}
        @media print {{
            body {{ margin: 0; }}
            .no-print {{ display: none; }}
            h1 {{ page-break-after: avoid; }}
            .summary {{ page-break-after: avoid; }}
            table {{ page-break-inside: auto; }}
            tr {{ page-break-inside: avoid; page-break-after: auto; }}
            thead {{ display: table-header-group; }}
        }}
    </style>
</head>
<body>"""

        # Add company name header if provided
        if company_name:
            html_out += f"""
    <div class="company-header">{html.escape(str(company_name))}</div>"""

        html_out += f"""
    <h1>{html.escape(str(title))}</h1>
    <div class="meta">Generated: {datetime.now(UTC).strftime("%Y-%m-%d %H:%M")}</div>
"""

        # Summary section
        if summary:
            html_out += f"""
    <div class="summary">
        <h2>Resumen</h2>
        <div class="stat">
            <div class="stat-value">{summary.total_records:,}</div>
            <div class="stat-label">Total de registros</div>
        </div>
    </div>
"""

        # Data table
        html_out += """
    <table>
        <thead>
            <tr>
"""
        for field in fields:
            html_out += (
                f"                <th>{html.escape(str(field.display_name))}</th>\n"
            )

        html_out += """            </tr>
        </thead>
        <tbody>
"""

        for row in data:
            html_out += "            <tr>\n"
            for field in fields:
                value = row.get(field.name)
                formatted = self._format_html_value(value)
                html_out += f"                <td>{formatted}</td>\n"
            html_out += "            </tr>\n"

        html_out += """        </tbody>
    </table>
</body>
</html>"""

        return html_out.encode("utf-8")

    def _html_to_pdf(self, html_content: bytes) -> bytes:
        """
        Convert HTML to PDF.

        Falls back to returning HTML if PDF generation not available.
        """
        # Try weasyprint
        try:
            from weasyprint import HTML  # type: ignore[import-untyped]

            return HTML(string=html_content.decode("utf-8")).write_pdf()  # type: ignore[no-any-return]
        except ImportError:
            logger.debug("weasyprint_not_installed_fallback")

        # Fallback: return HTML with PDF-like message
        # In production, you'd want to install weasyprint
        fallback_html = html_content.decode("utf-8")
        fallback_html = fallback_html.replace(
            "</body>",
            '<p style="color:red;margin-top:40px;">Note: Install weasyprint for PDF generation</p></body>',
        )
        return fallback_html.encode("utf-8")

    @staticmethod
    def _sanitize_formula(value: str) -> str:
        """Prevent Excel formula injection by prefixing dangerous chars with a single-quote."""
        if value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
            return "'" + value
        return value

    def _format_cell_value(self, value: Any) -> Any:
        """Format a value for Excel cells."""
        if value is None:
            return ""
        if isinstance(value, UUID):
            return str(value)
        if isinstance(value, bool):
            return "S\u00ed" if value else "No"
        if isinstance(value, str):
            return self._sanitize_formula(value)
        return value

    def _format_html_value(self, value: Any) -> str:
        """Format a value for HTML display (all output is HTML-escaped)."""
        if value is None:
            return "-"
        if isinstance(value, datetime):
            return html.escape(value.strftime("%Y-%m-%d %H:%M"))
        if isinstance(value, bool):
            return "✓" if value else "✗"
        if isinstance(value, UUID):
            return html.escape(str(value)[:8] + "...")
        return html.escape(str(value))


def get_reporter(session: AsyncSession) -> GenericReporter:
    """
    Get a GenericReporter instance.

    Args:
        session: SQLAlchemy async session

    Returns:
        GenericReporter instance
    """
    return GenericReporter(session)
