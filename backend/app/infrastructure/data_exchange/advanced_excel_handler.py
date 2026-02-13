# Copyright (c) 2025-2026 Sebastián Muñoz
# Licensed under the MIT License

"""
Advanced Excel Handler for enhanced Excel operations.

Provides additional Excel features including:
- Formulas (SUM, AVERAGE, COUNT, etc.)
- Charts (Bar, Pie, Line)
- Conditional formatting
- Data validation
- Named ranges
- Multiple sheets
- Protection settings
"""
# pyright: reportMissingImports=false, reportOptionalMemberAccess=false

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from enum import Enum
from typing import Any
from uuid import UUID

from app.infrastructure.observability.logging import get_logger

logger = get_logger(__name__)


# ============================================================================
# Check for openpyxl availability
# ============================================================================

OPENPYXL_AVAILABLE = False

# Initialize all openpyxl imports as None to avoid unbound errors
Workbook: Any = None
BarChart: Any = None
LineChart: Any = None
PieChart: Any = None
Reference: Any = None
DataLabelList: Any = None
ColorScaleRule: Any = None
FormulaRule: Any = None
IconSetRule: Any = None
Alignment: Any = None
Border: Any = None
Font: Any = None
PatternFill: Any = None
Side: Any = None
get_column_letter: Any = None
dataframe_to_rows: Any = None
DataValidation: Any = None

try:
    from openpyxl import Workbook  # type: ignore[import-not-found]
    from openpyxl.chart import (  # type: ignore[import-not-found]
        BarChart,
        LineChart,
        PieChart,
        Reference,
    )
    from openpyxl.formatting.rule import (  # type: ignore[import-not-found]
        ColorScaleRule,
        FormulaRule,
        IconSetRule,
    )
    from openpyxl.styles import (  # type: ignore[import-not-found]
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter  # type: ignore[import-not-found]
    from openpyxl.worksheet.datavalidation import (
        DataValidation,  # type: ignore[import-not-found]
    )

    OPENPYXL_AVAILABLE = True
except ImportError:
    logger.warning("openpyxl not installed. Advanced Excel features unavailable.")


def is_advanced_excel_available() -> bool:
    """Check if advanced Excel features are available."""
    return OPENPYXL_AVAILABLE


class ChartType(str, Enum):
    """Excel chart types."""

    BAR = "bar"
    COLUMN = "column"
    LINE = "line"
    PIE = "pie"


class FormatRule(str, Enum):
    """Conditional formatting rule types."""

    COLOR_SCALE = "color_scale"
    DATA_BAR = "data_bar"
    ICON_SET = "icon_set"
    FORMULA = "formula"


@dataclass
class ExcelChartConfig:
    """Configuration for Excel chart."""

    chart_type: ChartType
    title: str = ""
    data_range: str = ""  # e.g., "A1:B10"
    category_range: str = ""  # e.g., "A1:A10" for labels
    position: str = "E2"  # Cell where chart is placed
    width: int = 15  # Chart width in columns
    height: int = 10  # Chart height in rows
    style: int = 10  # Chart style (1-48)


@dataclass
class ConditionalFormatConfig:
    """Configuration for conditional formatting."""

    rule_type: FormatRule
    cell_range: str  # e.g., "B2:B100"

    # For color scale
    min_color: str = "F8696B"  # Red
    mid_color: str = "FFEB84"  # Yellow
    max_color: str = "63BE7B"  # Green

    # For formula rule
    formula: str = ""
    fill_color: str = "FFEB84"
    font_color: str = "000000"


@dataclass
class DataValidationConfig:
    """Configuration for data validation (dropdown, etc.)."""

    cell_range: str  # e.g., "C2:C100"
    validation_type: str = "list"  # list, whole, decimal, date, time, textLength

    # For list validation
    options: list[str] = field(default_factory=list)

    # For numeric validation
    min_value: float | None = None
    max_value: float | None = None

    # UI options
    show_dropdown: bool = True
    allow_blank: bool = True
    error_title: str = "Invalid Input"
    error_message: str = "Please enter a valid value"


@dataclass
class FormulaColumn:
    """Configuration for a formula column."""

    column_letter: str  # e.g., "F"
    header: str  # Column header
    formula_template: str  # e.g., "=SUM(B{row}:E{row})"
    number_format: str = "General"  # e.g., "#,##0.00", "0%"


@dataclass
class ExcelSheetConfig:
    """Configuration for an Excel sheet."""

    name: str
    headers: list[str] = field(default_factory=list)
    data: list[list[Any]] = field(default_factory=list)

    # Optional features
    formulas: list[FormulaColumn] = field(default_factory=list)
    charts: list[ExcelChartConfig] = field(default_factory=list)
    conditional_formats: list[ConditionalFormatConfig] = field(default_factory=list)
    validations: list[DataValidationConfig] = field(default_factory=list)

    # Layout options
    freeze_panes: str | None = "A2"  # Freeze first row
    auto_filter: bool = True
    column_widths: dict[str, int] = field(default_factory=dict)  # e.g., {"A": 20}

    # Protection
    protected: bool = False
    password: str | None = None


class AdvancedExcelHandler:
    """
    Advanced Excel handler with enhanced features.

    Provides charts, formulas, conditional formatting,
    and data validation capabilities.
    """

    # Default styling
    HEADER_FILL = "1a56db"  # Blue
    HEADER_FONT_COLOR = "ffffff"
    ALTERNATE_ROW_FILL = "f9fafb"
    BORDER_COLOR = "e5e7eb"

    def __init__(self):
        """Initialize advanced Excel handler."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel support. "
                "Install with: pip install openpyxl"
            )

    def create_workbook(
        self,
        sheets: list[ExcelSheetConfig],
        title: str = "Report",
    ) -> bytes:
        """
        Create an Excel workbook with multiple sheets.

        Args:
            sheets: List of sheet configurations
            title: Workbook title

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)

        for sheet_config in sheets:
            ws = wb.create_sheet(title=sheet_config.name[:31])  # Max 31 chars
            self._populate_sheet(ws, sheet_config)

        # Set document properties
        wb.properties.title = title
        wb.properties.creator = "FastAPI Enterprise"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read()

    def _populate_sheet(self, ws: Any, config: ExcelSheetConfig) -> None:
        """Populate a worksheet with data and features."""

        # Styles
        header_font = Font(bold=True, color=self.HEADER_FONT_COLOR)
        header_fill = PatternFill(
            start_color=self.HEADER_FILL, end_color=self.HEADER_FILL, fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin", color=self.BORDER_COLOR),
            right=Side(style="thin", color=self.BORDER_COLOR),
            top=Side(style="thin", color=self.BORDER_COLOR),
            bottom=Side(style="thin", color=self.BORDER_COLOR),
        )
        alt_fill = PatternFill(
            start_color=self.ALTERNATE_ROW_FILL,
            end_color=self.ALTERNATE_ROW_FILL,
            fill_type="solid",
        )

        # Write headers
        for col_idx, header in enumerate(config.headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Add formula columns headers
        for formula_col in config.formulas:
            col_idx = ord(formula_col.column_letter.upper()) - ord("A") + 1
            cell = ws.cell(row=1, column=col_idx, value=formula_col.header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data rows
        for row_idx, row_data in enumerate(config.data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(
                    row=row_idx, column=col_idx, value=self._format_value(value)
                )
                cell.border = border

                # Alternate row coloring
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Add formula columns
        for formula_col in config.formulas:
            col_idx = ord(formula_col.column_letter.upper()) - ord("A") + 1
            for row_idx in range(2, len(config.data) + 2):
                formula = formula_col.formula_template.replace("{row}", str(row_idx))
                cell = ws.cell(row=row_idx, column=col_idx, value=formula)
                cell.border = border
                cell.number_format = formula_col.number_format

                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Add summary formulas at bottom
        if config.data:
            summary_row = len(config.data) + 3
            ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)

            # Add SUM, AVERAGE, COUNT for numeric columns
            for col_idx in range(2, len(config.headers) + 1):
                # Check if column appears numeric
                col_letter = get_column_letter(col_idx)
                data_range = f"{col_letter}2:{col_letter}{len(config.data) + 1}"

                ws.cell(
                    row=summary_row, column=col_idx, value=f"=SUM({data_range})"
                ).font = Font(bold=True)
                ws.cell(
                    row=summary_row + 1, column=col_idx, value=f"=AVERAGE({data_range})"
                )
                ws.cell(
                    row=summary_row + 2, column=col_idx, value=f"=COUNT({data_range})"
                )

            # Labels
            ws.cell(row=summary_row, column=1, value="Total")
            ws.cell(row=summary_row + 1, column=1, value="Average")
            ws.cell(row=summary_row + 2, column=1, value="Count")

        # Set column widths
        for col_letter, width in config.column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Auto-width columns not specified
        for col_idx in range(1, len(config.headers) + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter not in config.column_widths:
                max_length = len(str(config.headers[col_idx - 1]))
                for row_data in config.data[:100]:  # Sample first 100 rows
                    if col_idx <= len(row_data):
                        max_length = max(max_length, len(str(row_data[col_idx - 1])))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Freeze panes
        if config.freeze_panes:
            ws.freeze_panes = config.freeze_panes

        # Auto filter
        if config.auto_filter and config.headers:
            last_col = get_column_letter(len(config.headers))
            ws.auto_filter.ref = f"A1:{last_col}1"

        # Add charts
        for chart_config in config.charts:
            self._add_chart(ws, chart_config)

        # Add conditional formatting
        for cf_config in config.conditional_formats:
            self._add_conditional_format(ws, cf_config)

        # Add data validation
        for dv_config in config.validations:
            self._add_data_validation(ws, dv_config)

        # Sheet protection
        if config.protected:
            ws.protection.sheet = True
            if config.password:
                ws.protection.password = config.password

    @staticmethod
    def _sanitize_formula(value: str) -> str:
        """Prevent Excel formula injection by prefixing dangerous chars with a single-quote."""
        if value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
            return "'" + value
        return value

    def _format_value(self, value: Any) -> Any:
        """Format value for Excel cell."""
        if isinstance(value, UUID):
            return str(value)
        if isinstance(value, (date, datetime)):
            return value
        if isinstance(value, str):
            return self._sanitize_formula(value)
        return value

    def _add_chart(self, ws: Any, config: ExcelChartConfig) -> None:
        """Add a chart to the worksheet."""
        try:
            # Parse data range
            if not config.data_range:
                return

            # Create chart based on type
            if config.chart_type == ChartType.PIE:
                chart = PieChart()
            elif config.chart_type == ChartType.LINE:
                chart = LineChart()
            else:
                chart = BarChart()
                if config.chart_type == ChartType.COLUMN:
                    chart.type = "col"
                else:
                    chart.type = "bar"

            chart.title = config.title
            chart.style = config.style

            # Parse range (e.g., "A1:B10")
            parts = config.data_range.split(":")
            if len(parts) == 2:
                start = parts[0]
                end = parts[1]

                # Get min/max row and column
                from openpyxl.utils import (
                    column_index_from_string,  # type: ignore[import-not-found]
                )

                start_col = column_index_from_string(
                    "".join(c for c in start if c.isalpha())
                )
                start_row = int("".join(c for c in start if c.isdigit()))
                end_col = column_index_from_string(
                    "".join(c for c in end if c.isalpha())
                )
                end_row = int("".join(c for c in end if c.isdigit()))

                data = Reference(
                    ws,
                    min_col=start_col,
                    min_row=start_row,
                    max_col=end_col,
                    max_row=end_row,
                )
                chart.add_data(data, titles_from_data=True)

                # Add category labels if specified
                if config.category_range:
                    cat_parts = config.category_range.split(":")
                    if len(cat_parts) == 2:
                        cat_start = cat_parts[0]
                        cat_end = cat_parts[1]
                        cat_col = column_index_from_string(
                            "".join(c for c in cat_start if c.isalpha())
                        )
                        cat_start_row = int(
                            "".join(c for c in cat_start if c.isdigit())
                        )
                        cat_end_row = int("".join(c for c in cat_end if c.isdigit()))

                        cats = Reference(
                            ws,
                            min_col=cat_col,
                            min_row=cat_start_row,
                            max_row=cat_end_row,
                        )
                        chart.set_categories(cats)

            # Size
            chart.width = config.width
            chart.height = config.height

            # Add to worksheet
            ws.add_chart(chart, config.position)

        except Exception as e:
            logger.warning("excel_chart_failed", error_type=type(e).__name__)

    def _add_conditional_format(self, ws: Any, config: ConditionalFormatConfig) -> None:
        """Add conditional formatting rule."""
        try:
            if config.rule_type == FormatRule.COLOR_SCALE:
                rule = ColorScaleRule(
                    start_type="min",
                    start_color=config.min_color,
                    mid_type="percentile",
                    mid_value=50,
                    mid_color=config.mid_color,
                    end_type="max",
                    end_color=config.max_color,
                )
                ws.conditional_formatting.add(config.cell_range, rule)

            elif config.rule_type == FormatRule.FORMULA and config.formula:
                rule = FormulaRule(
                    formula=[config.formula],
                    fill=PatternFill(
                        start_color=config.fill_color,
                        end_color=config.fill_color,
                        fill_type="solid",
                    ),
                    font=Font(color=config.font_color),
                )
                ws.conditional_formatting.add(config.cell_range, rule)

            elif config.rule_type == FormatRule.ICON_SET:
                rule = IconSetRule(
                    "3TrafficLights1",
                    "percent",
                    [0, 33, 67],
                    showValue=None,
                    percent=None,
                    reverse=None,
                )
                ws.conditional_formatting.add(config.cell_range, rule)

        except Exception as e:
            logger.warning("excel_conditional_format_failed", error_type=type(e).__name__)

    def _add_data_validation(self, ws: Any, config: DataValidationConfig) -> None:
        """Add data validation to cells."""
        try:
            if config.validation_type == "list" and config.options:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(config.options)}"',
                    showDropDown=not config.show_dropdown,
                    allow_blank=config.allow_blank,
                )
            elif config.validation_type in ("whole", "decimal"):
                formula1 = (
                    str(config.min_value) if config.min_value is not None else "0"
                )
                formula2 = (
                    str(config.max_value) if config.max_value is not None else "999999"
                )
                dv = DataValidation(
                    type=config.validation_type,
                    operator="between",
                    formula1=formula1,
                    formula2=formula2,
                    allow_blank=config.allow_blank,
                )
            else:
                return

            dv.error = config.error_message
            dv.errorTitle = config.error_title
            dv.add(config.cell_range)
            ws.add_data_validation(dv)

        except Exception as e:
            logger.warning("excel_data_validation_failed", error_type=type(e).__name__)

    def create_report_with_summary(
        self,
        title: str,
        headers: list[str],
        data: list[list[Any]],
        summary_data: dict[str, Any] | None = None,
        include_charts: bool = True,
    ) -> bytes:
        """
        Create a report Excel file with data and summary sheet.

        Args:
            title: Report title
            headers: Column headers
            data: Data rows
            summary_data: Optional summary statistics
            include_charts: Include charts in summary sheet

        Returns:
            Excel file as bytes
        """
        sheets = []

        # Main data sheet
        data_sheet = ExcelSheetConfig(
            name="Data",
            headers=headers,
            data=data,
            freeze_panes="A2",
            auto_filter=True,
        )

        # Add conditional formatting for numeric columns
        for col_idx, _header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            # Simple heuristic: if first data value is numeric, add color scale
            if data and col_idx <= len(data[0]):
                first_val = data[0][col_idx - 1]
                if isinstance(first_val, (int, float)):
                    data_sheet.conditional_formats.append(
                        ConditionalFormatConfig(
                            rule_type=FormatRule.COLOR_SCALE,
                            cell_range=f"{col_letter}2:{col_letter}{len(data) + 1}",
                        )
                    )

        sheets.append(data_sheet)

        # Summary sheet
        if summary_data:
            summary_headers = ["Metric", "Value"]
            summary_rows = [[k, v] for k, v in summary_data.items()]

            summary_sheet = ExcelSheetConfig(
                name="Summary",
                headers=summary_headers,
                data=summary_rows,
                column_widths={"A": 30, "B": 20},
            )

            # Add chart if we have numeric summary data
            if include_charts:
                numeric_items = [
                    (k, v)
                    for k, v in summary_data.items()
                    if isinstance(v, (int, float))
                ]
                if len(numeric_items) >= 2:
                    summary_sheet.charts.append(
                        ExcelChartConfig(
                            chart_type=ChartType.BAR,
                            title="Summary Overview",
                            data_range=f"A2:B{len(summary_rows) + 1}",
                            position="D2",
                        )
                    )

            sheets.append(summary_sheet)

        return self.create_workbook(sheets, title)


# ============================================================================
# Convenience Functions
# ============================================================================


def create_excel_report(
    title: str,
    headers: list[str],
    data: list[list[Any]],
    summary: dict[str, Any] | None = None,
) -> bytes:
    """
    Create an Excel report with optional summary.

    Args:
        title: Report title
        headers: Column headers
        data: Data rows
        summary: Optional summary statistics

    Returns:
        Excel file as bytes
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel support. Install with: pip install openpyxl"
        )

    handler = AdvancedExcelHandler()
    return handler.create_report_with_summary(title, headers, data, summary)


def create_multi_sheet_excel(
    sheets: list[ExcelSheetConfig], title: str = "Report"
) -> bytes:
    """
    Create a multi-sheet Excel workbook.

    Args:
        sheets: List of sheet configurations
        title: Workbook title

    Returns:
        Excel file as bytes
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel support. Install with: pip install openpyxl"
        )

    handler = AdvancedExcelHandler()
    return handler.create_workbook(sheets, title)
